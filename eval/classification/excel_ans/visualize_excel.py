import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 读取Excel文件
file_path = 'bias.xlsx'
df = pd.read_excel(file_path)

# 加载工作簿和工作表
workbook = load_workbook(file_path)
sheet = workbook.active

# 基准值为第二行数据，列从第二列开始
base_values = df.iloc[0, 1:]  # 第二行的数据
columns = df.columns[1:]  # 跳过第一列列名

# 每列计算最大变化范围
max_changes = {}
for col in columns:
    col_values = df[col].iloc[1:]  # 从第三行开始
    if base_values[col] != 0:
        max_change = max(abs((col_values - base_values[col]) / base_values[col]))
    else:
        max_change = max(abs(col_values))
    max_changes[col] = max_change if max_change > 0 else 1

# 定义颜色填充函数
def get_color(value, base_value, max_change):
    if base_value == 0:
        return None  # 避免除以0
    change = (value - base_value) / base_value  # 计算百分比变化
    normalized_change = abs(change) / max_change  # 归一化变化到 [0, 1]
    intensity = int(normalized_change * 255)  # 映射到颜色强度 [0, 255]
    if change > 0:
        # 增加 - 绿色
        return PatternFill(start_color=f'00{intensity:02X}FF00', end_color=f'00{intensity:02X}FF00', fill_type="solid")
    elif change < 0:
        # 减少 - 红色
        return PatternFill(start_color=f'00FF{intensity:02X}00', end_color=f'00FF{intensity:02X}00', fill_type="solid")
    else:
        # 无变化
        return None

# 应用颜色到工作表
for row_idx in range(3, sheet.max_row + 1):  # 从第三行开始（跳过行名）
    for col_idx in range(2, sheet.max_column + 1):  # 从第二列开始（跳过列名）
        col_name = sheet.cell(row=1, column=col_idx).value  # 获取列名
        current_value = sheet.cell(row=row_idx, column=col_idx).value
        base_value = sheet.cell(row=2, column=col_idx).value  # 第二行基准值
        if isinstance(current_value, (int, float)) and isinstance(base_value, (int, float)):
            max_change = max_changes[col_name]
            color = get_color(current_value, base_value, max_change)
            if color:
                sheet.cell(row=row_idx, column=col_idx).fill = color

# 保存修改后的Excel文件
output_file_path = 'bias_colored_corrected_final.xlsx'
workbook.save(output_file_path)
print(f"修正后的Excel文件已保存为 {output_file_path}")
